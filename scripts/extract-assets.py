import json
import re
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RESOURCE_DIR = ROOT / "已有资源"
DATA_DIR = ROOT / "data"

DIMENSIONS = [
    ("income_drive", "收入目标感"),
    ("growth_mode", "成长模式"),
    ("recognition_source", "认同感来源"),
    ("social_energy", "社交消耗度"),
    ("boundary_clarity", "边界清晰度"),
    ("recognition_logic", "认同逻辑"),
    ("pressure_rhythm", "压力节奏"),
    ("power_execution", "权力与执行"),
    ("thinking_granularity", "思维颗粒度"),
    ("information_processing", "信息处理"),
    ("drive_logic", "驱动逻辑"),
    ("feedback_cycle", "反馈周期"),
    ("environment_dependence", "环境空间依赖"),
    ("moral_threshold", "道德阈值"),
]

REVERSE_KEYWORDS = [
    "反转",
    "修正",
    "抵消",
    "更喜欢把一个已经掌握",
    "不在乎我的名片",
    "不需要任何人的赞赏",
    "多写一份报告",
    "严格区分",
    "工作就是工作",
    "成果导向",
    "不需要任何人拍马屁",
    "特别讨厌最后一刻",
    "不给我明确的指示",
    "现有的规则",
    "只说结论",
    "纸上谈兵",
    "数据证明",
    "看不到明显的进度",
    "10 年后的远大目标",
    "固定工位",
    "绝不会",
    "潜规则",
]

QUESTION_KEEP_INDEXES = {
    "income_drive": [1, 2, 3, 4, 6, 10],
    "growth_mode": [1, 2, 3, 4, 6, 10],
    "recognition_source": [1, 2, 3, 4, 6, 10],
    "social_energy": [1, 2, 3, 4, 5, 10],
    "boundary_clarity": [1, 2, 3, 4, 6, 10],
    "recognition_logic": [1, 2, 3, 4, 6, 10],
    "pressure_rhythm": [1, 2, 3, 4, 10],
    "power_execution": [1, 2, 3, 4, 10],
    "thinking_granularity": [1, 2, 3, 4, 10],
    "information_processing": [1, 2, 3, 4, 10],
    "drive_logic": [1, 2, 3, 4, 10],
    "feedback_cycle": [1, 2, 3, 4, 10],
    "environment_dependence": [1, 2, 3, 4, 10],
    "moral_threshold": [1, 2, 3, 4, 10],
}


def clean_question(raw: str) -> str:
    text = re.sub(r"^\([^)]*\)\s*", "", raw.strip())
    text = text.replace(" / 问：", " ").replace("问：", "")
    text = text.replace(
        "你会选择录取 A 吗？",
        "你更倾向录取 A 还是 B？",
    )
    return re.sub(r"\s+", " ", text)


def response_mode(dimension_id: str, question_no: int, text: str) -> str:
    if dimension_id == "drive_logic" and question_no == 8:
        return "choiceAB"
    if "感到" in text or "会觉得" in text:
        return "feeling"
    if "愿意" in text or "会选择" in text or "会接受" in text or "会主动" in text:
        return "action"
    return "agreement"


def extract_questionnaire():
    doc = Document(RESOURCE_DIR / "问题收集R1.docx")
    questions = []
    for dimension_index, table in enumerate(doc.tables):
        dimension_id, dimension_name = DIMENSIONS[dimension_index]
        core_logic = table.rows[0].cells[0].text.strip()
        question_no = 1
        for row in table.rows[1:]:
            raw = row.cells[-1].text.strip()
            if not raw or raw.isdigit() or raw.startswith("问："):
                continue
            if question_no not in QUESTION_KEEP_INDEXES[dimension_id]:
                question_no += 1
                continue
            reverse = any(keyword in raw for keyword in REVERSE_KEYWORDS)
            questions.append(
                {
                    "id": f"{dimension_id}_{question_no}",
                    "dimension": dimension_id,
                    "dimensionName": dimension_name,
                    "text": clean_question(raw),
                    "reverse": reverse,
                    "responseMode": response_mode(dimension_id, question_no, raw),
                    "source": "问题收集R1.docx",
                }
            )
            question_no += 1
    return {
        "scale": [
            {"value": 1, "label": "非常不同意"},
            {"value": 2, "label": "不太同意"},
            {"value": 3, "label": "说不清"},
            {"value": 4, "label": "比较同意"},
            {"value": 5, "label": "非常同意"},
        ],
        "dimensions": [
            {"id": dimension_id, "name": name, "order": index + 1}
            for index, (dimension_id, name) in enumerate(DIMENSIONS)
        ],
        "questions": questions,
    }


def extract_dimension_notes():
    wb = load_workbook(RESOURCE_DIR / "职业评估维度.xlsx", data_only=True)
    ws = wb.worksheets[0]
    notes = {}
    for row in ws.iter_rows(values_only=True):
        values = [cell for cell in row[:6]]
        if values[1] in [name for _, name in DIMENSIONS]:
            notes[values[1]] = {
                "block": values[0] or "",
                "low": values[2] or "",
                "medium": values[3] or "",
                "high": values[4] or "",
                "note": values[5] or "",
            }
    return notes


def sentence(name, band, note):
    tendency = note.get(band, "")
    if band == "low":
        return f"{name}偏低时，你更倾向于{tendency}。这不是优劣判断，而是在职业选择中要尊重的能量偏好。"
    if band == "high":
        return f"{name}偏高时，你更容易表现出{tendency}。当环境允许你按这种方式工作时，动力和效率会更稳定。"
    return f"{name}处在中间区间时，你通常能在不同工作方式之间切换，但需要更清晰的阶段目标来减少摇摆。"


def make_analysis_library(notes):
    library = []
    band_titles = {"low": "低分倾向", "medium": "平衡区间", "high": "高分倾向"}
    for dimension_id, name in DIMENSIONS:
        note = notes.get(name, {})
        for band in ["low", "medium", "high"]:
            library.append(
                {
                    "dimension": dimension_id,
                    "dimensionName": name,
                    "scoreBand": band,
                    "title": f"{name}：{band_titles[band]}",
                    "baseAnalysis": sentence(name, band, note),
                    "strengths": [
                        "能更快识别让自己稳定发挥的工作条件",
                        "在合适环境中更容易形成清晰的判断标准",
                        "可以把个人偏好转化为岗位筛选条件",
                    ],
                    "risks": [
                        "如果长期进入相反环境，容易把不适配误认为自己能力不足",
                        "在压力下可能放大单一倾向，忽略现实约束",
                    ],
                    "suggestions": [
                        "把这个维度写成求职筛选清单，而不是抽象性格标签",
                        "用 2-3 周小实验验证自己在不同环境中的真实消耗",
                        "和收入、成长、反馈周期一起综合判断，而不是单点决策",
                    ],
                    "suitableEnvironments": [
                        f"允许{name}偏好被看见的团队",
                        "目标、反馈和协作边界都相对清楚的岗位",
                    ],
                    "avoidEnvironments": [
                        f"长期要求你压抑{name}真实倾向的岗位",
                        "只看短期结果、缺少沟通和复盘机制的环境",
                    ],
                }
            )
    return library


CAREER_RULES = [
    ("增长产品与商业分析", ["growth_mode:high", "information_processing:medium", "drive_logic:high"], "适合把复杂信息拆成增长机会，用数据和用户反馈持续迭代。", ["增长产品经理", "商业分析师", "用户增长运营"], "先做一个真实产品的增长复盘，再补足 SQL、数据看板和实验设计能力。"),
    ("内容策略与品牌策划", ["recognition_source:high", "drive_logic:medium", "feedback_cycle:high"], "适合在表达、共鸣和传播反馈中获得动力。", ["内容策划", "品牌策略", "新媒体主编"], "沉淀 3 个主题账号案例，验证选题、转化和复盘能力。"),
    ("B2B 销售与客户成功", ["income_drive:high", "social_energy:high", "recognition_logic:high"], "适合高互动、高反馈、结果清晰的商业场景。", ["大客户销售", "客户成功经理", "解决方案顾问"], "从小客单价产品练习需求挖掘和复盘，逐步进入复杂交易。"),
    ("研究分析与咨询", ["thinking_granularity:high", "information_processing:high", "pressure_rhythm:medium"], "适合深度调研、结构化输出和严谨推理。", ["行业研究员", "战略咨询顾问", "政策研究"], "选择一个行业做 30 页研究报告，找业内人士校正判断。"),
    ("项目管理与组织推进", ["power_execution:medium", "pressure_rhythm:high", "boundary_clarity:medium"], "适合把混乱任务拆成节奏、责任和交付。", ["项目经理", "运营 PMO", "交付经理"], "从跨部门小项目开始，练习里程碑、风险台账和复盘。"),
    ("专业技术深耕", ["growth_mode:high", "thinking_granularity:high", "social_energy:low"], "适合在深度专注中累积壁垒。", ["软件工程师", "数据工程师", "算法应用工程师"], "选一个可展示项目，持续迭代 90 天，形成作品集。"),
    ("人力组织发展", ["recognition_logic:high", "drive_logic:medium", "social_energy:medium"], "适合在人成长、团队协作和组织机制中发挥影响。", ["HRBP", "组织发展顾问", "培训发展经理"], "从访谈、诊断和工作坊设计开始积累方法论。"),
    ("创业与新业务探索", ["income_drive:high", "power_execution:high", "growth_mode:high"], "适合在不确定中做判断、承担结果并快速学习。", ["创始人助理", "新业务负责人", "创业合伙人"], "用最小可行产品验证需求，先控制现金流和试错成本。"),
    ("财务风控与合规", ["moral_threshold:high", "thinking_granularity:high", "information_processing:high"], "适合在规则、细节和风险边界中建立专业价值。", ["风控经理", "合规专员", "审计"], "补齐法规框架和案例库，练习风险识别报告。"),
    ("空间运营与线下管理", ["environment_dependence:high", "power_execution:medium", "social_energy:medium"], "适合依托秩序、现场和流程做稳定交付。", ["门店运营", "园区运营", "行政运营"], "从一个线下流程优化项目切入，记录成本、效率和体验指标。"),
    ("自由职业与远程服务", ["boundary_clarity:high", "environment_dependence:low", "feedback_cycle:high"], "适合用清晰边界和快速反馈管理个人产能。", ["自由撰稿人", "独立顾问", "远程运营"], "先用副业订单验证交付节奏，再设计报价和客户筛选标准。"),
    ("教育培训与知识产品", ["growth_mode:high", "recognition_logic:high", "feedback_cycle:medium"], "适合把经验结构化，并通过他人成长获得成就感。", ["课程研发", "职业讲师", "学习顾问"], "打磨一个 60 分钟公开课，用反馈优化课程闭环。"),
]


def make_career_rules():
    return [
        {
            "trackName": name,
            "conditions": conditions,
            "reason": reason,
            "suitableRoles": roles,
            "entryPath": entry,
            "risks": "需要避免只被单一优势吸引，应同步验证收入、节奏和长期成长空间。",
            "validationMethod": "用 30-90 天低成本项目、访谈或兼职任务做真实验证。",
        }
        for name, conditions, reason, roles, entry in CAREER_RULES
    ]


COMBINATIONS = [
    ("growth_mode", "information_processing", "thinking_granularity", "深度学习型建构者", "你适合把复杂知识变成可复用方法，而不是只做零散执行。"),
    ("income_drive", "pressure_rhythm", "power_execution", "高目标攻坚型", "高目标会激活你，但需要明确风险底线，避免用持续透支换短期结果。"),
    ("social_energy", "recognition_logic", "drive_logic", "关系场域影响者", "你更容易在需要理解人、协调人、说服人的场景中发挥。"),
    ("boundary_clarity", "feedback_cycle", "environment_dependence", "自我管理型交付者", "清楚边界和反馈节奏会显著影响你的稳定产出。"),
    ("moral_threshold", "thinking_granularity", "information_processing", "规则风险守门人", "你适合在标准、合规和严谨验证中建立专业可信度。"),
    ("recognition_source", "income_drive", "feedback_cycle", "外部成就驱动者", "可见成果会推动你前进，但要避免把全部价值感交给外部评价。"),
    ("social_energy", "boundary_clarity", "pressure_rhythm", "能量节奏敏感型", "你需要同时管理互动密度、休息边界和任务波峰。"),
    ("power_execution", "growth_mode", "drive_logic", "开创式问题解决者", "你适合在模糊目标下找到路径，但需要复盘机制减少试错成本。"),
    ("recognition_logic", "moral_threshold", "income_drive", "价值边界谈判者", "高回报机会出现时，你需要提前定义不可交换的职业底线。"),
    ("environment_dependence", "social_energy", "thinking_granularity", "专注空间依赖者", "空间和干扰程度会直接影响你的细节质量与决策耐心。"),
    ("feedback_cycle", "growth_mode", "pressure_rhythm", "短周期迭代者", "你适合在快速反馈中成长，长期项目需要主动拆阶段成果。"),
    ("information_processing", "drive_logic", "power_execution", "理性决策推进者", "你擅长用信息和模型降低不确定性，再推动团队行动。"),
    ("income_drive", "boundary_clarity", "environment_dependence", "高回报边界管理者", "你可以追求更高收入，但需要评估工作对生活边界的占用。"),
    ("recognition_source", "social_energy", "feedback_cycle", "曝光反馈成长者", "外部反馈能给你动力，内容、销售、培训等场景可重点验证。"),
    ("moral_threshold", "recognition_logic", "drive_logic", "温和原则型", "你适合有清晰价值观的团队，面对灰色操作会有明显消耗。"),
    ("thinking_granularity", "feedback_cycle", "pressure_rhythm", "质量节奏平衡者", "你要在细节质量和交付速度之间建立自己的标准线。"),
    ("power_execution", "boundary_clarity", "social_energy", "授权协作管理者", "明确授权与沟通边界时，你更容易带动多人交付。"),
    ("growth_mode", "recognition_source", "income_drive", "上升通道追求者", "你需要能同时满足成长、回报和可见成就的赛道。"),
    ("information_processing", "environment_dependence", "boundary_clarity", "结构化独立工作者", "你适合清楚规则和独立思考空间并存的工作方式。"),
    ("drive_logic", "recognition_logic", "moral_threshold", "复杂人事判断者", "你在涉及人、情理和边界的决策中需要保留书面依据。"),
]


def make_combination_rules():
    return [
        {
            "id": f"combo_{index + 1}",
            "name": name,
            "conditions": [f"{a}:high", f"{b}:high", f"{c}:high"],
            "insight": insight,
            "suitableTracks": ["增长产品与商业分析", "研究分析与咨询", "项目管理与组织推进"],
            "risk": "组合优势如果缺少场景验证，可能停留在自我感觉良好。",
            "suggestion": "挑一个最贴近当前背景的方向，用小项目或访谈验证真实适配度。",
        }
        for index, (a, b, c, name, insight) in enumerate(COMBINATIONS)
    ]


def main():
    DATA_DIR.mkdir(exist_ok=True)
    notes = extract_dimension_notes()
    files = {
        "questionnaire.json": extract_questionnaire(),
        "analysis-library.json": make_analysis_library(notes),
        "career-rules.json": make_career_rules(),
        "combination-rules.json": make_combination_rules(),
    }
    for filename, payload in files.items():
        (DATA_DIR / filename).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        print(filename)


if __name__ == "__main__":
    main()
